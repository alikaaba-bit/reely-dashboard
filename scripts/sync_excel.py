#!/usr/bin/env python3
"""
Excel to Supabase Sync Script
Reads Reely client data from Excel and syncs to Supabase
"""

import os
import sys
from datetime import datetime
from typing import List, Dict

try:
    import openpyxl
    from supabase import create_client
except ImportError:
    print("Installing dependencies...")
    os.system("pip3 install openpyxl supabase")
    import openpyxl
    from supabase import create_client

# Configuration
EXCEL_PATH = os.environ.get('EXCEL_PATH', '/Users/ali/.openclaw/media/inbound/file_23---4d58603b-e29a-4a31-b0a7-8ae4bf36414d.xlsx')
SUPABASE_URL = os.environ.get('NEXT_PUBLIC_SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')

def parse_clients_from_excel(filepath: str) -> List[Dict]:
    """Extract client data from Excel Clients tab"""
    print(f"📊 Reading Excel: {filepath}")
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    if 'Clients' not in wb.sheetnames:
        print("❌ No 'Clients' tab found")
        return []
    
    ws = wb['Clients']
    clients = []
    
    # Read from row 3 (skip headers)
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and row[0] != 'Company Name':
            company = row[0]
            status = row[1] or 'Unknown'
            monthly_rate = row[2] or 0
            additional = row[3] or ''
            
            # Parse additional billing
            additional_amount = 0
            if additional and '$' in str(additional):
                import re
                match = re.search(r'\$([0-9,]+)', str(additional))
                if match:
                    additional_amount = int(match.group(1).replace(',', ''))
            
            clients.append({
                'company_name': company,
                'status': status,
                'monthly_rate': monthly_rate,
                'additional_billing': additional_amount,
                'total_monthly': monthly_rate + additional_amount,
                'synced_at': datetime.now().isoformat()
            })
    
    print(f"✅ Found {len(clients)} clients")
    return clients

def sync_to_supabase(clients: List[Dict]) -> bool:
    """Sync client data to Supabase"""
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("❌ Missing Supabase credentials")
        print("Set: NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY")
        return False
    
    try:
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        
        # Calculate MRR totals
        active_clients = [c for c in clients if c['status'] == 'Active']
        total_mrr = sum(c['total_monthly'] for c in active_clients)
        
        # Update mrr_metrics table
        today = datetime.now().strftime('%Y-%m-%d')
        
        mrr_data = {
            'date': today,
            'mrr': total_mrr,
            'active_clients': len(active_clients),
            'new_clients': 0,  # Would need historical data
            'churned_clients': 0,  # Would need historical data
        }
        
        supabase.table('mrr_metrics').upsert(mrr_data, on_conflict='date').execute()
        print(f"✅ Synced MRR: ${total_mrr:,} from {len(active_clients)} clients")
        
        # Update scorecard goals if they exist
        # This would pull from Q1 2026, Jan, Feb, March tabs
        
        return True
        
    except Exception as e:
        print(f"❌ Supabase error: {e}")
        return False

def main():
    """Main sync function"""
    print("🚀 Reely Excel Sync Starting...")
    print("=" * 50)
    
    # Check if file exists
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Excel file not found: {EXCEL_PATH}")
        sys.exit(1)
    
    # Parse clients
    clients = parse_clients_from_excel(EXCEL_PATH)
    
    if not clients:
        print("❌ No clients found in Excel")
        sys.exit(1)
    
    # Print summary
    print("\n📋 Client Summary:")
    print("-" * 50)
    for c in clients[:10]:  # Show first 10
        print(f"  {c['company_name']:<25} ${c['total_monthly']:>8,}/mo ({c['status']})")
    
    if len(clients) > 10:
        print(f"  ... and {len(clients) - 10} more")
    
    # Sync to Supabase
    print("\n☁️  Syncing to Supabase...")
    if sync_to_supabase(clients):
        print("\n✅ Sync complete!")
    else:
        print("\n⚠️  Sync failed (check credentials)")
        # Still show the data we would have synced
        active = [c for c in clients if c['status'] == 'Active']
        total = sum(c['total_monthly'] for c in active)
        print(f"\n📊 Would sync: ${total:,} from {len(active)} active clients")

if __name__ == '__main__':
    main()
